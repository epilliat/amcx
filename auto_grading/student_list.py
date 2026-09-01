"""Charge la liste des étudiants depuis EXAM_2026/Liste_etudiants_avec_numeros(1).xlsx
et fournit le matching id_4_chiffres -> étudiant canonique + fallback fuzzy par nom.
"""

from __future__ import annotations

import difflib
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from config import load_config, resolve_path

ROOT = Path(__file__).resolve().parent


@dataclass(frozen=True)
class Student:
    id: str      # 5 chiffres
    nom: str
    prenom: str

    @property
    def last4(self) -> str:
        return self.id[-4:]

    @property
    def full(self) -> str:
        return f"{self.nom} {self.prenom}"


def _norm(s: str) -> str:
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-zA-Z]+", " ", s).strip().lower()
    return s


def xlsx_header(path) -> list[str]:
    """Retourne les noms de colonnes (1re ligne) d'un fichier xlsx."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    header = next(ws.iter_rows(values_only=True), ())
    wb.close()
    return [str(c) for c in header if c is not None]


def load_students() -> list[Student]:
    """Charge la liste depuis le fichier xlsx + colonnes définis dans config.json.

    Retourne une liste vide si aucun xlsx n'a été fourni (projet vierge).
    """
    cfg = load_config()
    raw = (cfg.get("student_xlsx") or "").strip()
    if not raw:
        return []
    path = resolve_path(raw)
    if not path.exists() or path.is_dir():
        return []
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c) if c is not None else "" for c in rows[0]]

    def col_idx(name: str, fallback: int) -> int:
        return header.index(name) if name in header else fallback

    i_id = col_idx(cfg["xlsx_id_col"], 0)
    i_nom = col_idx(cfg["xlsx_nom_col"], 1)
    i_prenom = col_idx(cfg["xlsx_prenom_col"], 2) if cfg["xlsx_prenom_col"] else None

    out = []
    for row in rows[1:]:
        if i_id >= len(row) or row[i_id] is None:
            continue
        raw_id = row[i_id]
        # Une cellule numérique perd ses zéros de tête dès Excel (`01234` → 1234).
        # Une cellule texte les conserve : on ne passe par int() que pour retirer
        # le `.0` des flottants.
        if isinstance(raw_id, float) and raw_id.is_integer():
            sid = str(int(raw_id))
        elif isinstance(raw_id, int):
            sid = str(raw_id)
        else:
            sid = str(raw_id).strip()
        nom = str(row[i_nom] or "") if i_nom < len(row) else ""
        prenom = str(row[i_prenom] or "") if (i_prenom is not None and i_prenom < len(row)) else ""
        out.append(Student(id=sid, nom=nom, prenom=prenom))
    return _pad_leading_zeros(out)


def _pad_leading_zeros(students: list[Student]) -> list[Student]:
    """Restaure les zéros de tête perdus par les cellules numériques du xlsx.

    Les identifiants d'un établissement sont de largeur fixe. Si une nette
    majorité (≥ 80 %) des ids numériques fait L caractères, ceux qui sont plus
    courts ont perdu leurs zéros de tête au stockage → on les complète. Sans ça,
    `by_full_id` (donc les `_student_override` posés à la relecture) ne matche
    plus, et `last4` est décalé.
    """
    digits = [s.id for s in students if s.id.isdigit()]
    if len(digits) < 2:
        return students
    widths: dict[int, int] = {}
    for d in digits:
        widths[len(d)] = widths.get(len(d), 0) + 1
    width, count = max(widths.items(), key=lambda kv: kv[1])
    if count / len(digits) < 0.8:
        return students        # largeurs hétérogènes : on ne devine rien
    return [
        Student(id=s.id.zfill(width), nom=s.nom, prenom=s.prenom)
        if (s.id.isdigit() and len(s.id) < width) else s
        for s in students
    ]


class StudentMatcher:
    def __init__(self) -> None:
        self.students = load_students()
        # index last4 (ignore les collisions éventuelles avec un autre xlsx)
        self._by_last4 = {}
        for s in self.students:
            if len(s.id) >= 4 and s.id[-4:].isdigit():
                self._by_last4.setdefault(s.id[-4:], s)
        self._by_id = {s.id: s for s in self.students}  # match par id complet
        # index de noms normalisés pour fuzzy match. Deux homonymes partagent
        # la même clé : le second écrasait le premier sans un mot. On garde le
        # premier et on signale, sinon un match par nom peut désigner l'autre.
        self._norm_to_student = {}
        self._homonyms: list[str] = []
        for s in self.students:
            key = _norm(s.full)
            if key in self._norm_to_student:
                self._homonyms.append(s.full)
                continue
            self._norm_to_student[key] = s
        if self._homonyms:
            print(f"⚠ liste étudiants : {len(self._homonyms)} homonyme(s) — "
                  f"le match par nom est ambigu pour : "
                  f"{', '.join(sorted(set(self._homonyms))[:5])}"
                  + (" …" if len(set(self._homonyms)) > 5 else ""))
        self._norm_keys = list(self._norm_to_student.keys())

    def by_full_id(self, sid: str) -> Student | None:
        """Match par identifiant complet (utilisé pour les overrides utilisateur)."""
        return self._by_id.get(str(sid))

    def by_id(self, id_4: str) -> Student | None:
        """id_4: chaîne lue par le modèle (potentiellement avec '?'). Retourne None si pas match unique."""
        if not id_4 or len(id_4) != 4 or not id_4.isdigit():
            return None
        return self._by_last4.get(id_4)

    def by_name(self, raw_name: str, cutoff: float = 0.7) -> tuple[Student | None, float]:
        """Fuzzy match d'un nom manuscrit contre la liste. Retourne (match, score)."""
        q = _norm(raw_name)
        if not q:
            return None, 0.0
        matches = difflib.get_close_matches(q, self._norm_keys, n=1, cutoff=cutoff)
        if not matches:
            return None, 0.0
        m = matches[0]
        score = difflib.SequenceMatcher(None, q, m).ratio()
        return self._norm_to_student[m], score

    def candidates(self, raw_name: str, n: int = 5,
                   cutoff: float = 0.4) -> list[tuple[Student, float]]:
        """Top-n étudiants les plus proches d'un nom (pour résolution manuelle)."""
        q = _norm(raw_name)
        if not q:
            return []
        out = []
        for m in difflib.get_close_matches(q, self._norm_keys, n=n, cutoff=cutoff):
            score = difflib.SequenceMatcher(None, q, m).ratio()
            out.append((self._norm_to_student[m], score))
        return out

    def resolve(self, id_4: str, raw_name: str) -> dict:
        """Stratégie complète:
        - matchID si possible
        - sinon fallback nom
        - sinon flag pour relecture
        Retourne dict: matched_student (Student|None), method, score, flag
        """
        s = self.by_id(id_4)
        if s is not None:
            return {"matched": s, "method": "id", "score": 1.0, "flag": ""}
        s, score = self.by_name(raw_name)
        if s is not None:
            return {"matched": s, "method": "name_fuzzy", "score": score,
                    "flag": f"ID '{id_4}' invalide, fallback nom (score={score:.2f})"}
        return {"matched": None, "method": "none", "score": 0.0,
                "flag": f"AUCUN MATCH (id={id_4!r}, name={raw_name!r})"}


if __name__ == "__main__":
    m = StudentMatcher()
    print(f"Chargé: {len(m.students)} étudiants")
    print("Quelques exemples:")
    for s in m.students[:3]:
        print(f"  {s.id} (last4={s.last4})  {s.nom} {s.prenom}")
    # tests
    print()
    print("Test by_id('2995'):", m.by_id("2995"))
    # exemples noms cités par l'utilisateur
    for raw in ["LASSALLE Gwenaël", "Lasalle Gwenael", "Ledo Mathis", "Delanzy Romane",
                "NADIN Sarah", "LASFARGUES Amandine"]:
        s, score = m.by_name(raw)
        print(f"  by_name({raw!r}) -> {s} (score={score:.2f})")
