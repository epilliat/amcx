"""Génère le fichier de notes au format exigé par la scolarité.

Reprend la structure EXACTE de note_Introduction_aux_tests_statistiques_(1ASTA10).xlsx
(colonne clé_verif, noms, métadonnées, colonne Ecrit) et remplit deux colonnes :
  - QCM         (colonne F) : score QCM brut sur 32
  - note finale (colonne G) : note finale agrégée

Les valeurs viennent de compte_rendu/notes.csv (généré par le bouton
« Sauvegarder le compte rendu » du dashboard). Le fichier modèle d'origine
n'est jamais modifié — on écrit un nouveau xlsx.

Usage:
    python auto_grading/export_scolarite.py [chemin_sortie.xlsx]
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent          # installation : pour les imports
sys.path.insert(0, str(ROOT))
import config  # noqa: E402
from student_list import StudentMatcher  # noqa: E402

_DATA = config.project_root()                    # projet actif : pour les données
NOTES_CSV = _DATA / "compte_rendu" / "notes.csv"
DEFAULT_OUT = _DATA / "compte_rendu" / "notes_scolarite.xlsx"


def template_path() -> Path | None:
    """Modèle xlsx scolarité (config `export_template_xlsx`), ou None si non configuré."""
    p = config.load_config().get("export_template_xlsx", "")
    return config.resolve_path(p) if p else None

DATA_START_ROW = 6   # 1re ligne de données (openpyxl, 1-indexé)
NAME_COL = 2         # colonne B : nom prénom
QCM_COL = 6          # colonne F : QCM
FINALE_COL = 7       # colonne G : note finale


def _num(v):
    """Convertit une chaîne csv en float, ou None si vide/illisible."""
    if v is None or str(v).strip() == "":
        return None
    try:
        return float(str(v).strip())
    except ValueError:
        return None


def main(out_path: Path) -> None:
    template = template_path()
    if template is None:
        sys.exit("Aucun modèle scolarité configuré "
                 "(clé config 'export_template_xlsx'). Export non disponible.")
    if not template.exists():
        sys.exit(f"Modèle introuvable : {template}")
    if not NOTES_CSV.exists():
        sys.exit(f"Introuvable : {NOTES_CSV}\n"
                 "→ clique « Sauvegarder le compte rendu » dans le dashboard d'abord.")

    # 1. notes.csv → {id_canonique: (qcm_brut_sur_32, note_finale)}
    by_id: dict[str, tuple] = {}
    with open(NOTES_CSV, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            sid = (row.get("id_canonique") or "").strip()
            if sid:
                by_id[sid] = (_num(row.get("QCM_brut_sur_32")),
                              _num(row.get("note_finale")))

    # 2. matcher pour relier un nom du fichier scolarité → étudiant → id
    matcher = StudentMatcher()

    # 3. ouvrir le modèle (copie en mémoire), remplir QCM + note finale
    wb = openpyxl.load_workbook(template)
    ws = wb.active
    ws.cell(row=1, column=FINALE_COL, value="note finale")   # en-tête colonne G

    n_filled = n_missing = 0
    for r in range(DATA_START_ROW, ws.max_row + 1):
        name = ws.cell(row=r, column=NAME_COL).value
        if name is None or not str(name).strip():
            continue
        student, _score = matcher.by_name(str(name))
        rec = by_id.get(student.id) if student is not None else None
        if rec is None:
            n_missing += 1
            print(f"  ⚠ pas de note pour : {name}")
            continue
        qcm, finale = rec
        ws.cell(row=r, column=QCM_COL, value=qcm)
        ws.cell(row=r, column=FINALE_COL, value=finale)
        n_filled += 1

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"\n✓ {out_path}")
    print(f"  {n_filled} étudiants remplis ; {n_missing} sans correspondance.")


if __name__ == "__main__":
    out = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else DEFAULT_OUT
    main(out)
